from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import TimeoutException
from urllib.parse import quote
from pyquery import PyQuery as pq
import os
import openpyxl
import random
import time

browser = webdriver.Chrome(r'D:\pydata\chromedriver\chromedriver.exe')
base_url = 'https://s.taobao.com/search?q='
keywords = '3到10岁儿童玩具'
url = base_url + quote(keywords)
wait = WebDriverWait(browser, 20)

page_max = 100


def log_out(browser):
    login_switch = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'div.login-switch i#J_Quick2Static')))
    login_switch.click()
    weibo_login = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.weibo-login')))
    weibo_login.click()
    username = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.inp.username input')))
    password = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.inp.password input')))
    username.send_keys('888')
    password.send_keys('888')
    submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'span[node-type="submitStates"]')))
    submit.click()


def get_page(page):
    print('正在打印 %d 页' % page)
    try:
        if page == 1:
            browser.get(url)
            if '手机扫码，安全登录' in browser.page_source:
                log_out(browser)
        else:
            input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'input.J_Input')))
            submit = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'span.J_Submit')))
            input.clear()
            input.send_keys(page)
            submit.click()
        wait.until(EC.text_to_be_present_in_element((By.CSS_SELECTOR, 'div#mainsrp-pager ul.items li.item.active span'),
                                                    str(page)))
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, 'div.m-itemlist')))
        if browser.page_source != None:
            return browser
        else:
            get_page(page)
    except TimeoutException as e:
        get_page(page)


def get_products(browser):
    html = browser.page_source
    doc = pq(html)
    for item in doc('#mainsrp-itemlist .items .item').items():
        image = item.find('.pic .img').attr('data-src')
        price = item.find('.price').text().replace('\n', '')
        deal = item.find('.deal-cnt').text()
        title = item.find('.title').text()
        shop = item.find('.shop').text()
        location = item.find('.location').text().replace(' ', '')
        yield [image, price, deal, title, shop, location]


def save(out):
    filename = 'taobao_' + keywords + '.xlsx'
    if not os.path.exists(filename):
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet(index=0, title=keywords)
        sheet.append(['图片', '价格', '成交人数', '商品', '店铺', '地点'])
        workbook.save(filename)
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[keywords]
    for row in out:
        print(row)
        sheet.append(row)
    workbook.save(filename)


def main():
    for page in range(1, page_max + 1):
        browser = get_page(page)
        out = get_products(browser)
        save(out)
        time.sleep(random.randint(1, 5))


if __name__ == '__main__':
    main()